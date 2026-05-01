from openpyxl import Workbook, load_workbook

# --- Zapis do pliku Excel ---
wb = Workbook()
ws = wb.active # Aktywny arkusz
ws.title = "finance"
ws.append(["nazwa", "wartość"]) # Dodaj nagłówek
ws.append(["czynsz", 2500])
ws.append(["jedzenie", 400])

ws["B4"] = "=SUM(B2:B3)"

wb.save("finanse.xlsx")
